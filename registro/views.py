import datetime
import json
import re
from abc import ABC, abstractmethod
from typing import List, Dict, Any
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError, models, transaction
from django.db.models import Q, Sum
from django.forms import modelformset_factory, formset_factory
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView, FormView
from sympy import sympify, Symbol
from weasyprint import CSS, HTML
from weasyprint.text.fonts import FontConfiguration

from nomencladores.models import EstadoAccion, TipoAccion, TipoMoneda, TipoPresupuesto, EstadoPresupuesto, \
    TipoIndicador, Escenario, Sector
from registro.services import FormulaCalculatorService, ResultadoIndicadorService, VariationCalculatorService, \
    ChartDataService, BreadcrumbBuilder, StatisticsCalculatorService, \
    InsightGeneratorService, RankingCalculatorService, MetaProgressService, PresupuestoAnalyticsService, \
    PresupuestoForecastService, PresupuestoVisualizationService, PresupuestoBenchmarkService, \
    PresupuestoIndicadorAnalyzer, ReportePDFService
from registro.forms import DocumentoForm, AccionForm, PresupuestoPlanificadoForm, PresupuestoEjecutadoForm, \
    IndicadorForm, VariableIndicadorForm, ResultadoVariableForm, ResultadoIndicadorForm
from registro.models import Accion, Documento, PresupuestoPlanificado, PresupuestoEjecutado, VariableIndicador, \
    Indicador, ResultadoIndicador, ResultadoVariable
from registro.utils import data_chart_donut, data_chart_line


# Create your views here.

class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

    def get(self, request, *args, **kwargs):
        # messages.success(request, 'Hola Bienvenidos al registro de acciones de adaptación')
        if not request.user.last_login:
            request.session['show_tour'] = True
        return super().get(request, *args, **kwargs)

    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     context['show_tour'] = self.request.session.get('show_tour', False)
    #     context['title_html'] = 'Inicio'
    #
    #     if 'show_tour' in self.request.session:
    #         del self.request.session['show_tour']
    #
    #     return context

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Filtrar acciones del usuario (o todas si es superuser)
        if user.is_superuser:
            acciones = Accion.objects.filter(publicado=True)
        else:
            acciones = Accion.objects.filter(user=user, publicado=True)

        # ============ RESUMEN GENERAL ============
        # Obtener todos los indicadores de las acciones
        indicadores_ids = []
        for accion in acciones:
            indicadores_ids.extend(accion.indicadores.values_list('id', flat=True))

        total_indicadores = len(set(indicadores_ids))  # Usar set para evitar duplicados
        indicadores_con_metas = Indicador.objects.filter(
            id__in=indicadores_ids,
            meta_valor__isnull=False
        ).count()

        context['resumen_general'] = {
            'total_acciones': acciones.count(),
            'acciones_activas': acciones.filter(
                estado_accion__nombre__icontains='ejecucion'
            ).count(),
            'total_indicadores': total_indicadores,
            'indicadores_con_metas': indicadores_con_metas,
        }

        # ============ PRESUPUESTOS POR MONEDA ============
        presupuestos_por_moneda = []
        for moneda in TipoMoneda.objects.filter(estado=True):
            # Planificado total - iterar sobre acciones
            total_planificado = 0
            total_ejecutado = 0

            for accion in acciones:
                # Sumar presupuestos planificados de esta acción con esta moneda
                presupuestos_accion = accion.presupuestos_planificados.filter(
                    tipo_moneda=moneda
                )
                for pp in presupuestos_accion:
                    total_planificado += pp.monto

                    # Sumar ejecutados de este presupuesto planificado
                    ejecutado_pp = pp.presupuestos_ejecutados.aggregate(
                        total=Sum('monto')
                    )['total'] or 0
                    total_ejecutado += ejecutado_pp

            # Cálculos
            restante = total_planificado - total_ejecutado
            porcentaje_ejecutado = (
                (total_ejecutado / total_planificado * 100)
                if total_planificado > 0 else 0
            )

            # Estado del semáforo
            if porcentaje_ejecutado < 50:
                estado = 'danger'
            elif porcentaje_ejecutado < 80:
                estado = 'warning'
            else:
                estado = 'success'

            presupuestos_por_moneda.append({
                'moneda': moneda.nombre,
                'total_planificado': round(total_planificado, 2),
                'total_ejecutado': round(total_ejecutado, 2),
                'restante': round(restante, 2),
                'porcentaje_ejecutado': round(porcentaje_ejecutado, 2),
                'estado': estado
            })

        context['presupuestos_por_moneda'] = presupuestos_por_moneda

        # ============ INDICADORES CRÍTICOS (Alertas) ============
        indicadores_criticos = []

        for accion in acciones:
            for indicador in accion.indicadores.all():
                alertas = []
                nivel_riesgo = 'bajo'

            # Verificar si tiene resultados
            resultados = indicador.resultados.all().order_by('-fecha')
            if not resultados.exists():
                alertas.append('Sin mediciones registradas')
                nivel_riesgo = 'alto'
            else:
                # Verificar última medición
                ultima_medicion = resultados.first()
                dias_sin_medir = (
                        timezone.now() - ultima_medicion.fecha
                ).days

                # Alerta por falta de mediciones recientes
                if dias_sin_medir > 90:
                    alertas.append(f'Sin mediciones hace {dias_sin_medir} días')
                    nivel_riesgo = 'medio'

                # Verificar progreso de meta
                if indicador.meta_valor:
                    progreso = indicador.calcular_progreso_meta()
                    if progreso:
                        porcentaje = progreso['progreso_porcentaje']

                        # Calcular tiempo transcurrido vs progreso
                        if indicador.meta_fecha_limite:
                            dias_totales = (
                                    indicador.meta_fecha_limite -
                                    accion.fecha_inicio
                            ).days
                            dias_transcurridos = (
                                    timezone.now().date() -
                                    accion.fecha_inicio
                            ).days
                            porcentaje_tiempo = (
                                dias_transcurridos / dias_totales * 100
                                if dias_totales > 0 else 0
                            )

                            # Si el progreso está muy por debajo del tiempo
                            if porcentaje < (porcentaje_tiempo - 20):
                                alertas.append(
                                    f'Progreso ({porcentaje:.1f}%) '
                                    f'por debajo del tiempo esperado '
                                    f'({porcentaje_tiempo:.1f}%)'
                                )
                                nivel_riesgo = 'alto'

                # Verificar tendencia negativa
                if resultados.count() >= 3:
                    ultimos_3 = list(resultados[:3])
                    if (ultimos_3[0].valor < ultimos_3[1].valor <
                            ultimos_3[2].valor):
                        if indicador.direccion_optima == 'aumentar':
                            alertas.append('Tendencia descendente detectada')
                            nivel_riesgo = 'medio'

            # Solo agregar si hay alertas
            if alertas:
                indicadores_criticos.append({
                    'indicador': indicador,
                    'accion': accion,  # Ya tenemos la acción del loop
                    'alertas': alertas,
                    'nivel_riesgo': nivel_riesgo,
                    'url': f'/accion/{accion.id}/indicadores/'
                           f'{indicador.id}/resultado/comportamiento/'
                })

            # Ordenar por nivel de riesgo
        orden_riesgo = {'alto': 0, 'medio': 1, 'bajo': 2}
        indicadores_criticos.sort(
            key=lambda x: orden_riesgo.get(x['nivel_riesgo'], 3)
        )

        context['indicadores_criticos'] = indicadores_criticos[:10]

        # ============ ACCIONES POR ESTADO ============
        acciones_por_estado = []
        for estado in EstadoAccion.objects.all().order_by('orden'):
            count = acciones.filter(estado_accion=estado).count()
            acciones_por_estado.append({
                'estado': estado.nombre,
                'count': count,
                'porcentaje': (
                    count / acciones.count() * 100
                    if acciones.count() > 0 else 0
                )
            })

        context['acciones_por_estado'] = acciones_por_estado

        # ============ TOP 5 ACCIONES CON MEJOR DESEMPEÑO ============
        acciones_con_score = []

        for accion in acciones:
            indicadores = accion.indicadores.all()
            if not indicadores.exists():
                continue

            score_total = 0
            indicadores_evaluados = 0

            for ind in indicadores:
                resultados = ind.resultados.all()
                if not resultados.exists():
                    continue

                indicadores_evaluados += 1

                # Score basado en:
                # 1. Progreso de meta (si existe)
                score_meta = 0
                if ind.meta_valor:
                    progreso = ind.calcular_progreso_meta()
                    if progreso:
                        score_meta = min(
                            progreso['progreso_porcentaje'], 100
                        )

                # 2. Consistencia de mediciones
                score_mediciones = min(
                    resultados.count() * 10, 100
                )

                # 3. Tendencia positiva
                score_tendencia = 50
                if resultados.count() >= 2:
                    variacion = ind.variacion_valor
                    if variacion and variacion.get('variacion_porcentual'):
                        var_pct = variacion['variacion_porcentual']
                        if ind.direccion_optima == 'aumentar':
                            score_tendencia = 50 + min(var_pct, 50)
                        else:
                            score_tendencia = 50 + min(-var_pct, 50)

                # Score promedio del indicador
                score_indicador = (
                        score_meta * 0.5 +
                        score_mediciones * 0.3 +
                        score_tendencia * 0.2
                )
                score_total += score_indicador

            if indicadores_evaluados > 0:
                score_final = score_total / indicadores_evaluados
                acciones_con_score.append({
                    'accion': accion,
                    'score': round(score_final, 2),
                    'total_indicadores': indicadores.count(),
                    'sector': accion.sector.nombre
                })

        # Ordenar por score y tomar top 5
        acciones_con_score.sort(key=lambda x: x['score'], reverse=True)
        context['top_acciones'] = acciones_con_score[:5]

        # ============ ANÁLISIS POR SECTOR ============
        sectores_data = []
        for sector in Sector.objects.all():
            acciones_sector = acciones.filter(sector=sector)

            if not acciones_sector.exists():
                continue

            # ✅ CORRECTO: Contar indicadores iterando sobre acciones
            total_indicadores = 0
            for accion in acciones_sector:
                total_indicadores += accion.indicadores.count()

            # ✅ CORRECTO: Presupuesto iterando sobre acciones
            presupuesto_sector = 0
            for accion in acciones_sector:
                for pp in accion.presupuestos_planificados.all():
                    presupuesto_sector += pp.monto

            sectores_data.append({
                'sector': sector.nombre,
                'total_acciones': acciones_sector.count(),
                'total_indicadores': total_indicadores,
                'presupuesto_total': presupuesto_sector
            })

        context['sectores_data'] = sectores_data

        # ============ PRÓXIMOS VENCIMIENTOS ============
        hoy = timezone.now()
        proximos_30_dias = hoy + datetime.timedelta(days=30)

        metas_proximas = []
        for accion in acciones:
            for indicador in accion.indicadores.filter(
                    meta_fecha_limite__isnull=False,
                    meta_fecha_limite__gte=hoy,
                    meta_fecha_limite__lte=proximos_30_dias
            ):
                progreso = indicador.calcular_progreso_meta()
                dias_restantes = (indicador.meta_fecha_limite - hoy).days

                # Nivel de urgencia
                if dias_restantes <= 7:
                    urgencia = 'alta'
                elif dias_restantes <= 15:
                    urgencia = 'media'
                else:
                    urgencia = 'baja'

                metas_proximas.append({
                    'indicador': indicador,
                    'accion': accion,
                    'dias_restantes': dias_restantes,
                    'progreso': progreso['progreso_porcentaje'] if progreso else 0,
                    'urgencia': urgencia
                })

        context['metas_proximas'] = metas_proximas

        # ============ METADATA ============
        context.update({
            'title_html': 'Resumen',
            'title_head': 'Panel de Control',
            'ultima_actualizacion': timezone.now()
        })

        return context


#############Action#############
class ActionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Accion
    template_name = 'action/list_action.html'
    permission_required = 'registro.view_accion'
    context_object_name = 'actions'

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect('registro:home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Listado acciones',
            'title_head': 'Acciones de adaptacion',
            'estados_accion': EstadoAccion.objects.all(),
            'accion': True,
            'crear_url': reverse_lazy('registro:registrar_accion'),
            'breadcrumbs': breadcrumbs,

        })
        return context


class ActionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Accion
    form_class = AccionForm
    template_name = 'action/create_action.html'
    permission_required = 'registro.add_accion'
    success_url = reverse_lazy('registro:home')

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect('registro:lista_accion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Nuevo', 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Nueva acción',
            'title_head': 'Nueva acción',
            'url_cancel': reverse_lazy('registro:lista_accion'),
            'form_docs': DocumentoForm(),
            'tipo_acciones': TipoAccion.objects.all().order_by('orden'),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 1
        })
        return context

    def form_valid(self, form):
        try:
            self.object = form.save(commit=False)
            self.object.estado_accion = EstadoAccion.objects.filter(orden=1).first()
            self.object.user = self.request.user
            self.object.save()

            # Procesar documentos adjuntos
            documentos = self.request.FILES.getlist('documentos')
            for documento in documentos:
                doc = Documento.objects.create(
                    documento=documento,
                    nombre=documento.name
                )
                self.object.documentos.add(doc)

            messages.success(self.request, 'La acción se ha registrado correctamente')
            return HttpResponseRedirect(reverse('registro:lista_presupuesto_planificado', args=[self.object.id]))

        except IntegrityError:
            messages.error(self.request, 'Ha ocurrido un error. Contacte con su administrador.')
            return self.form_invalid(form)


class ActionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Accion
    form_class = AccionForm
    template_name = 'action/create_action.html'
    permission_required = 'app.change_accion'  # Reemplaza 'app' con tu nombre de app
    permission_denied_message = 'Usted no tiene los privilegios necesarios para esta operación.'

    def get_success_url(self):
        return reverse('registro:lista_presupuesto_planificado', args=[self.object.id])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': self.object.id, 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Editar acción',
            'title_head': f'Editar acción [{self.object.id}]',
            'url_cancel': reverse('registro:lista_accion'),
            'form_docs': DocumentoForm(),
            'tipo_acciones': TipoAccion.objects.all().order_by('orden'),
            'selected_tipo_acciones': self.object.tipo_accion,
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 1
        })
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        form_docs = DocumentoForm(self.request.POST, self.request.FILES)

        if form_docs.is_valid():
            documentos = self.request.FILES.getlist('documentos')
            if documentos:
                for f in documentos:
                    doc = Documento.objects.create(
                        documento=f,
                        nombre=f.name
                    )
                    self.object.documentos.add(doc)

        messages.success(self.request, 'La acción se ha modificado correctamente')
        return HttpResponseRedirect(
            reverse('registro:lista_presupuesto_planificado', args=[self.object.id]))

    def handle_no_permission(self):
        messages.error(self.request, self.permission_denied_message)
        return redirect('registro:lista_accion')


class DetailtsActionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Accion
    template_name = 'action/detalle_accion.html'
    permission_required = 'app.view_accion'  # Reemplaza 'app' con tu nombre de app
    permission_denied_message = 'Usted no tiene los privilegios necesarios para esta operación.'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': self.object.id, 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Detalles acción',
            'title_head': f'Detalles de la acción [{self.object.id}]',
            'url_cancel': reverse('registro:lista_accion'),
            'breadcrumbs': breadcrumbs,
            'presupuestos_planificados': self.object.presupuestos_planificados.all(),
        })
        return context

    def handle_no_permission(self):
        messages.error(self.request, self.permission_denied_message)
        return redirect('registro:lista_accion')


@csrf_exempt
@login_required
@permission_required('registro.delete_accion', raise_exception=True)
def eliminar_accion(request, id_accion):
    if request.method == 'POST':
        try:
            accion = get_object_or_404(Accion, id=id_accion)
            accion.delete()
            return JsonResponse({'success': True, 'message': 'La acción ha sido eliminada correctamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error al eliminar: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)


#############Presupuesto#############
class PresupuestoPlanificadoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'registro.view_presupuestoplanificado'
    template_name = 'presupuesto/presupuesto_planificado.html'
    context_object_name = 'presupuestos_planificado'

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_accion',
                                args=[self.kwargs['id_accion']]))

    def get_queryset(self):
        self.accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])
        return self.accion.presupuestos_planificados.all().order_by('tipo_presupuesto__orden')

    def get(self, request, *args, **kwargs):
        # Redirigir si no hay presupuestos
        if self.get_queryset().count() == 0:
            return HttpResponseRedirect(
                reverse('registro:registrar_presupuesto_planificado',
                        args=[self.kwargs['id_accion']]))
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto', 'url': None, 'icon': None},
            {'name': 'Planificado', 'url': None, 'icon': None, 'active': True},
        ]
        tipos_monedas = TipoMoneda.objects.filter(estado=True)
        totales_presupuestos = []
        desglose_total = []

        for tm in tipos_monedas:
            presupuestos = self.accion.presupuestos_planificados.filter(
                tipo_moneda=tm).order_by('tipo_presupuesto__orden')

            if presupuestos:
                total = presupuestos.aggregate(total=models.Sum('monto'))['total']
                monto_ejecutado_total = 0
                desglose_planificado = []
                subtotales = []

                chart_donut_data = data_chart_donut()

                for pp in presupuestos:
                    subtotales.append(pp.monto)
                    monto_planificado = pp.monto
                    presupuestos_ejecutados = pp.presupuestos_ejecutados.all()
                    monto_ejecutado = sum(p.monto for p in presupuestos_ejecutados)

                    if monto_planificado > 0:
                        porcentaje_ejecucion = (monto_ejecutado / monto_planificado) * 100
                        restante = monto_planificado - monto_ejecutado
                        porcentaje_restante = (restante / monto_planificado) * 100
                    else:
                        porcentaje_ejecucion = 0
                        restante = 0
                        porcentaje_restante = 0

                    desglose_planificado.append({
                        'moneda': tm.nombre,
                        'planificado': monto_planificado,
                        'ejecutado': monto_ejecutado,
                        'porcentaje_ejecucion': porcentaje_ejecucion,
                        'restante': restante,
                        'porcentaje_restante': porcentaje_restante,
                        'total': total
                    })
                    monto_ejecutado_total += monto_ejecutado

                if monto_ejecutado_total > 0 and total:
                    porcentaje_ejecucion_total = (monto_ejecutado_total / total) * 100
                    restante_total = total - monto_ejecutado_total
                    porcentaje_restante_total = (restante_total / total) * 100
                else:
                    porcentaje_ejecucion_total = 0
                    restante_total = 0
                    porcentaje_restante_total = 0

                chart_donut_data['series'] = [total - monto_ejecutado_total, monto_ejecutado_total] if total else [0, 0]

                desglose_total.append({
                    'moneda': tm.sigla,
                    'desglose_planificado': desglose_planificado,
                    'porcentaje_ejecucion_total': porcentaje_ejecucion_total,
                    'monto_ejecutado_total': monto_ejecutado_total,
                    'monto_planificado_total': total,
                    'restante_total': restante_total,
                    'chart_donut_data': chart_donut_data,
                    'porcentaje_restante_total': porcentaje_restante_total
                })

                totales_presupuestos.append({
                    'moneda': tm.sigla,
                    'subtotales': subtotales,
                    'total': total
                })

        context.update({
            'title_html': 'Presupuesto acción',
            'title_head': f'Presupuestos planificados de la acción [{self.kwargs["id_accion"]}]',
            'url_cancel': reverse('registro:editar_accion', args=[self.kwargs['id_accion']]),
            'accion': self.accion,
            'crear_url': reverse('registro:registrar_presupuesto_planificado',
                                 args=[self.kwargs['id_accion']]),
            'totales_presupuestos': totales_presupuestos,
            'desglose_presupuesto': desglose_total,
            'tipo_presupuestos': TipoPresupuesto.objects.all(),
            'estado_presupuestos': EstadoPresupuesto.objects.all(),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 2,
            'next_url': reverse('registro:lista_indicador', args=[self.kwargs['id_accion']])
        })

        return context


class PresupuestoPlanificadoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    form_class = PresupuestoPlanificadoForm
    template_name = 'presupuesto/create_update_presupuesto_planificado.html'
    permission_required = 'app.add_presupuestoplanificado'  # Reemplaza 'app' con tu nombre de app

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accion = Accion.objects.filter(id=self.kwargs['id_accion']).first()

        # Determinar URL de cancelación
        url_cancelar = reverse('registro:editar_accion', args=[self.kwargs['id_accion']])
        if accion.presupuestos_planificados.all().count() > 0:
            url_cancelar = reverse('registro:lista_presupuesto_planificado',
                                   args=[self.kwargs['id_accion']])
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto',
             'url': reverse('registro:lista_presupuesto_planificado', args=[self.kwargs['id_accion']]), 'icon': None},
            {'name': 'Planificado', 'url': None, 'icon': None},
            {'name': 'Nuevo', 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Presupuesto acción',
            'title_head': f'Presupuestos planificados de la acción [{self.kwargs["id_accion"]}]',
            'url_cancel': url_cancelar,
            'form_docs': DocumentoForm(),
            'accion': accion,
            'tipo_presupuestos': TipoPresupuesto.objects.all(),
            'estado_presupuestos': EstadoPresupuesto.objects.all(),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 2
        })
        return context

    def form_valid(self, form):
        try:
            self.object = form.save()
            accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])
            accion.presupuestos_planificados.add(self.object)

            messages.success(self.request, 'El presupuesto planificado se ha registrado correctamente')
            return HttpResponseRedirect(
                reverse('registro:lista_presupuesto_planificado',
                        args=[self.kwargs['id_accion']]))
        except IntegrityError:
            messages.error(self.request, 'Ha ocurrido un error contacte con su administrador.')
            return self.form_invalid(form)

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_presupuesto_planificado',
                                args=[self.kwargs['id_accion']]))


class PresupuestoPlanificadoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    form_class = PresupuestoPlanificadoForm
    template_name = 'presupuesto/create_update_presupuesto_planificado.html'
    permission_required = 'app.change_presupuestoplanificado'  # Reemplaza 'app' con tu nombre de app

    def get_object(self, queryset=None):
        return get_object_or_404(PresupuestoPlanificado, id=self.kwargs['id_presupuesto'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto',
             'url': reverse('registro:lista_presupuesto_planificado', args=[self.kwargs['id_accion']]), 'icon': None},
            {'name': 'Planificado', 'url': None, 'icon': None},
            {'name': 'Editar', 'url': None, 'icon': None, 'active': True},
        ]
        context.update({
            'title_html': 'Presupuesto acción',
            'title_head': f'Presupuestos planificados de la acción [{self.kwargs["id_accion"]}]',
            'url_cancel': reverse('registro:lista_presupuesto_planificado',
                                  args=[self.kwargs['id_accion']]),
            'accion': accion,
            'selected_tipo_presupuesto': self.object.tipo_presupuesto,
            'tipo_presupuestos': TipoPresupuesto.objects.all(),
            'estado_presupuestos': EstadoPresupuesto.objects.all(),
            'presupuestos_planificado': accion.presupuestos_planificados.all().order_by('tipo_presupuesto__orden'),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 2
        })
        return context

    def form_valid(self, form):
        try:
            self.object = form.save()
            messages.success(self.request, 'El presupuesto planificado se ha modificado correctamente')
            return HttpResponseRedirect(
                reverse('registro:lista_presupuesto_planificado',
                        args=[self.kwargs['id_accion']]))
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error contacte con su administrador.')
            return self.form_invalid(form)

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_presupuesto_planificado',
                                args=[self.kwargs['id_accion']]))


@login_required
@permission_required('registro.delete_presupuestoplanificado', raise_exception=True)
def eliminar_presupuesto_planificado(request, id_accion, id_presupuesto):
    try:
        presupuesto_planificado = get_object_or_404(PresupuestoPlanificado, id=id_presupuesto)

        if request.POST:
            with transaction.atomic():
                presupuesto_planificado.presupuestos_ejecutados.all().delete()
                presupuesto_planificado.delete()
            messages.success(request, 'El presupuesto planificado se ha eliminado correctamente')
            return HttpResponseRedirect(
                reverse('registro:lista_presupuesto_planificado', args=[id_accion]))

    except PermissionDenied:
        messages.error(request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_presupuesto_planificado', args=[id_accion]))


#############PresupuestoEjecutado#############

class PresupuestoEjecutadoView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    form_class = PresupuestoEjecutadoForm
    template_name = 'presupuesto/presupuestos_ejecutados.html'
    permission_required = 'registro.add_presupuestoejecutado'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_accion = self.kwargs['id_accion']
        id_presupuesto = self.kwargs['id_presupuesto']
        accion = Accion.objects.filter(id=id_accion).first()
        presupuesto_planificado = PresupuestoPlanificado.objects.filter(id=id_presupuesto).first()
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto', 'url': reverse('registro:lista_presupuesto_planificado', args=[id_accion]),
             'icon': None},
            {'name': 'Planificado [' + str(id_presupuesto) + ']',
             'url': reverse('registro:editar_presupuesto_planificado', args=[id_accion, id_presupuesto]), 'icon': None},
            {'name': 'Nuevo', 'url': None, 'icon': None, 'active': True},
        ]

        context.update({
            'title_html': 'Presupuesto ejecutado',
            'title_head': f'Presupuestos ejecutados de la planificación [{id_presupuesto}] y la acción [{id_accion}]',
            'url_cancel': reverse('registro:lista_presupuesto_planificado', args=[id_accion]),
            'accion': accion,
            'presupuesto_planificado': presupuesto_planificado,
            'tipo_presupuestos': TipoPresupuesto.objects.all(),
            'estado_presupuestos': EstadoPresupuesto.objects.all(),
            'presupuestos_planificado': accion.presupuestos_planificados.all().order_by('-monto'),
            'presupuestos_ejecutados': presupuesto_planificado.presupuestos_ejecutados.all().order_by(
                'presupuestos_ejecutados__tipo_presupuesto__orden'),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 2.1
        })
        return context

    def form_valid(self, form):
        id_accion = self.kwargs['id_accion']
        id_presupuesto = self.kwargs['id_presupuesto']
        accion = get_object_or_404(Accion, id=id_accion)
        presupuesto_planificado = get_object_or_404(PresupuestoPlanificado, id=id_presupuesto)

        try:
            with transaction.atomic():
                obj = form.save()
                if accion.estado_accion.orden == 1:
                    accion.estado_accion.orden = accion.estado_accion.orden + 1
                    accion.save()
                presupuesto_planificado.presupuestos_ejecutados.add(obj)

            messages.success(self.request, 'El presupuesto ejecutado se ha registrado correctamente')
            return HttpResponseRedirect(
                reverse('registro:registrar_presupuesto_ejecutado',
                        args=[id_accion, id_presupuesto]))

        except Exception as e:
            messages.error(self.request, f'{str(e)} Ha ocurrido un error contacte con su administrador.')
            return self.form_invalid(form)

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_presupuesto_planificado',
                                args=[self.kwargs['id_accion']]))


class PresupuestoEjecutadoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = PresupuestoEjecutado
    form_class = PresupuestoEjecutadoForm
    template_name = 'presupuesto/presupuestos_ejecutados.html'
    permission_required = 'registro.change_presupuestoejecutado'

    def dispatch(self, request, *args, **kwargs):
        self.accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])
        self.presupuesto_planificado = get_object_or_404(
            PresupuestoPlanificado,
            id=self.kwargs['id_presupuesto']
        )
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return get_object_or_404(PresupuestoEjecutado, id=self.kwargs['id_ejecutado'])

    def get_success_url(self):
        messages.success(self.request, 'El presupuesto ejecutado se ha modificado correctamente')
        return reverse('registro:registrar_presupuesto_ejecutado',
                       args=[self.kwargs['id_accion'], self.kwargs['id_presupuesto']])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_accion = self.kwargs['id_accion']
        id_presupuesto = self.kwargs['id_presupuesto']
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto', 'url': reverse('registro:lista_presupuesto_planificado', args=[id_accion]),
             'icon': None},
            {'name': 'Planificado [' + str(self.kwargs['id_presupuesto']) + ']',
             'url': reverse('registro:editar_presupuesto_planificado', args=[id_accion, id_presupuesto]), 'icon': None},
            {'name': 'Editar', 'url': None, 'icon': None, 'active': True},
            {'name': 'Ejecutado [' + str(self.kwargs['id_ejecutado']) + ']', 'url': None, 'icon': None, 'active': True},
        ]

        totales_presupuestos = []
        tipos_monedas = TipoMoneda.objects.all()

        for tm in tipos_monedas:
            subtotales = []
            presupuestos = self.accion.presupuestos_planificados.filter(tipo_moneda=tm)
            total = presupuestos.aggregate(total=models.Sum('monto'))['total']
            for pm in presupuestos:
                subtotales.append(pm.monto)
            totales_presupuestos.append(
                {'moneda': tm.nombre, 'subtotales': subtotales, 'total': total})

        context.update({
            'title_html': 'Presupuesto ejecutado acción',
            'title_head': f'Editar presupuesto ejecutado de la planificación [{id_presupuesto}] y la acción [{id_accion}]',
            'url_cancel': reverse('registro:lista_presupuesto_planificado',
                                  args=[id_accion]),
            'accion': self.accion,
            'presupuesto_planificado': self.presupuesto_planificado,
            'totales_presupuestos': totales_presupuestos,
            'tipo_presupuestos': TipoPresupuesto.objects.all(),
            'estado_presupuestos': EstadoPresupuesto.objects.all(),
            'presupuestos_planificado': self.accion.presupuestos_planificados.all().order_by('-monto'),
            'presupuestos_ejecutados': self.presupuesto_planificado.presupuestos_ejecutados.all().order_by('-monto'),
            'show_menu_left': True,
            'breadcrumbs': breadcrumbs,
            'current_step': 2.1
        })
        return context

    def handle_no_permission(self):
        messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_presupuesto_planificado',
                                args=[self.kwargs['id_accion']]))


@login_required
@permission_required('registro.delete_presupuestoejecutado', raise_exception=True)
def eliminar_presupuesto_ejecutado(request, id_accion, id_presupuesto, id_ejecutado):
    try:
        presupuesto_planificado = PresupuestoPlanificado.objects.filter(id=id_presupuesto).first()
        presupuesto_ejecutado = PresupuestoEjecutado.objects.filter(id=id_ejecutado).first()

        if request.POST:
            presupuesto_ejecutado.delete()
            messages.success(request, 'El presupuesto ejecutado se ha eliminado correctamente')
            return HttpResponseRedirect(reverse('registro:registrar_presupuesto_ejecutado',
                                                args=[id_accion, presupuesto_planificado.id]))
    except PermissionDenied:
        messages.error(request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:registrar_presupuesto_ejecutado', args=[id_accion]))


###########Indicadores################
# class IndicadoresListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
#     permission_required = 'registro.view_indicador'
#     template_name = 'indicador/indicadores.html'
#     context_object_name = 'indicadores'
#
#     def handle_no_permission(self):
#         messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
#         return redirect(reverse('registro:lista_accion',
#                                 args=[self.kwargs['id_accion']]))
#
#     def get_queryset(self):
#         self.accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])
#         return self.accion.indicadores.all().order_by('nombre')
#
#     def get(self, request, *args, **kwargs):
#         if self.get_queryset().count() == 0:
#             return HttpResponseRedirect(
#                 reverse('registro:registrar_indicador',
#                         args=[self.kwargs['id_accion']]))
#         return super().get(request, *args, **kwargs)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         breadcrumbs = [
#             {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
#             {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
#             {'name': 'Indicadores', 'url': None, 'icon': None},
#         ]
#         context.update({
#             'title_html': f'Listado indicadores de {self.accion.nombre}',
#             'title_head': f'Indicadores de la acción [{self.kwargs["id_accion"]}]',
#             'cobeneficio': True,
#             'accion': self.accion,
#             'crear_url': reverse('registro:registrar_indicador', args=[self.kwargs['id_accion']]),
#             'show_menu_left': True,
#             'breadcrumbs': breadcrumbs,
#             'current_step': 3,
#             'next_url': '#'
#         })
#         return context
#
#
# class IndicadorCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
#     form_class = IndicadorForm
#     template_name = 'indicador/crear_editar_indicador_accion.html'
#     permission_required = 'registro.add_indicador'
#
#     def dispatch(self, request, *args, **kwargs):
#         self.id_accion = kwargs.get('id_accion')
#         self.accion = get_object_or_404(Accion, id=self.id_accion)
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_success_url(self):
#         return reverse('registro:lista_indicador', args=[self.id_accion])
#
#     def get_cancel_url(self):
#         if self.accion.indicadores.all().count() > 0:
#             return reverse('registro:lista_indicador', args=[self.id_accion])
#         return reverse('registro:lista_accion')
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         breadcrumbs = [
#             {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
#             {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
#             {'name': 'Indicadores', 'url': reverse('registro:lista_indicador', args=[self.id_accion]), 'icon': None},
#             {'name': 'Nuevo Indicador', 'url': None, 'icon': None},
#         ]
#         context.update({
#             'title_html': 'Indicador',
#             'title_head': f'Indicador de la acción [{self.id_accion}]',
#             'url_cancel': self.get_cancel_url(),
#             'form_variable_indicador': VariableIndicadorForm(),
#             'accion': self.accion,
#             'show_menu_left': True,
#             'indicadores': self.accion.indicadores.all(),
#             'tipo_indicador': TipoIndicador.objects.all(),
#             'breadcrumbs': breadcrumbs,
#             'current_step': 3.1,
#             'next_url': '#'
#         })
#         return context
#
#     def form_valid(self, form):
#         try:
#             with transaction.atomic():
#                 json_data = json.loads(self.request.POST.get('form_variables', '[]'))
#                 variables_formula = re.findall(r'\b[a-zA-z_]+\b', form.cleaned_data['formula'])
#                 variables = [js['variable'] for js in json_data]
#                 for vf in variables_formula:
#                     if vf not in variables:
#                         messages.error(self.request,
#                             f"La variable [{vf}] no aparece en la fórmula del indicador. Por favor revise que las variables esten bien descritas.")
#                         return HttpResponseRedirect(reverse('registro:lista_indicador', args=[self.id_accion]))
#                 obj = form.save()
#                 self.accion.indicadores.add(obj)
#                 for f in json_data:
#                     if f.get('nombre'):
#                         var, created = VariableIndicador.objects.get_or_create(nombre=f['nombre'])
#                         var.variable = f['variable']
#                         var.save()
#                         obj.variable_indicador.add(var)
#             messages.success(self.request, 'El indicador se ha registrado correctamente')
#             return HttpResponseRedirect(self.get_success_url())
#         except Exception as e:
#             messages.error(self.request, '%s Ha ocurrido un error contacte con su administrador.' % str(e))
#             return self.form_invalid(form)
#
#     def handle_no_permission(self):
#         messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
#         return redirect(reverse('registro:lista_indicador', args=[self.id_accion]))
#
#
# class IndicadorUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
#     model = Indicador
#     form_class = IndicadorForm
#     template_name = 'indicador/crear_editar_indicador_accion.html'
#     permission_required = 'registro.change_indicador'  # Cambiado a registro
#     permission_denied_message = 'Usted no tiene los privilegios necesarios para esta operación.'
#
#     def dispatch(self, request, *args, **kwargs):
#         self.id_accion = kwargs.get('id_accion')
#         self.accion = get_object_or_404(Accion, id=self.id_accion)
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_object(self, queryset=None):
#         # Obtener el objeto usando id_indicador en lugar de pk
#         id_indicador = self.kwargs.get('id_indicador')
#         return get_object_or_404(Indicador, id=id_indicador)
#
#     def get_success_url(self):
#         return reverse('registro:lista_indicador', args=[self.id_accion])
#
#     def get_cancel_url(self):
#         accion = get_object_or_404(Accion, id=self.id_accion)
#         if accion.indicadores.all().count() > 0:
#             return reverse('registro:lista_indicador', args=[self.id_accion])
#         return reverse('registro:lista_accion')
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         accion = get_object_or_404(Accion, id=self.id_accion)
#         VariableIndicadorFormSet = modelformset_factory(VariableIndicador, form=VariableIndicadorForm, extra=0, can_delete=True)
#         breadcrumbs = [
#             {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
#             {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
#             {'name': 'Indicadores', 'url': reverse('registro:lista_indicador', args=[self.id_accion]), 'icon': None},
#             {'name': 'Indicador', 'url': None, 'icon': None},
#         ]
#         context.update({
#             'title_html': 'Editar Indicador',
#             'title_head': f'Indicador de la acción [{self.id_accion}]',
#             'url_cancel': self.get_cancel_url(),
#             # 'form_variable_indicador': VariableIndicadorForm(),
#             'formset': VariableIndicadorFormSet(queryset=self.object.variable_indicador.all()),
#             'accion': self.accion,
#             'show_menu_left': True,
#             'indicadores': accion.indicadores.all(),
#             'tipo_indicador': TipoIndicador.objects.all(),
#             'breadcrumbs': breadcrumbs,
#             'current_step': 3.1,
#             'next_url': '#'
#         })
#         return context
#
#     def form_valid(self, form):
#         try:
#             with transaction.atomic():
#
#                 json_data = self._get_json_data()
#                 variables_formula = self._extract_variables_from_formula(form.cleaned_data['formula'])
#                 variables = self._extract_variables_from_json(json_data)
#                 for vf in variables_formula:
#                     if vf not in variables:
#                         messages.error(self.request,
#                                        f"La variable [{vf}] no aparece en la fórmula del indicador. Por favor revise que las variables esten bien descritas.")
#                         return HttpResponseRedirect(reverse('registro:lista_indicador', args=[self.id_accion]))
#
#                 obj = form.save()
#                 obj.variable_indicador.clear()
#
#                 if not self._save_related_variables(obj, json_data):
#                     return self._redirect_with_error(
#                         'registro:lista_indicador',
#                         self.id_accion,
#                         "Error al guardar las variables relacionadas"
#                     )
#
#             messages.success(self.request, 'El indicador se ha registrado correctamente')
#             return HttpResponseRedirect(self.get_success_url())
#         except Exception as e:
#             messages.error(self.request, '%s Ha ocurrido un error contacte con su administrador.' % str(e))
#             return self.form_invalid(form)
#
#     def _get_json_data(self):
#         json_string = self.request.POST.get('formset_variables', '[]')
#         return json.loads(json_string)
#
#     def _extract_variables_from_formula(self, formula):
#         """Extrae variables de la fórmula usando expresión regular"""
#         return re.findall(r'\b[a-zA-Z_][a-zA-Z0-9_]*\b', formula)
#
#     def _extract_variables_from_json(self, json_data):
#         """Extrae nombres de variables del JSON"""
#         return [item['variable'] for item in json_data if 'variable' in item]
#
#     def _save_related_variables(self, obj, json_data):
#         """Guarda las variables relacionadas de forma segura"""
#         try:
#             for item in json_data:
#                 if item.get('nombre'):
#                     variable, created = VariableIndicador.objects.get_or_create(
#                         nombre=item['nombre'],
#                         defaults={'variable': item['variable']}
#                     )
#
#                     # Actualizar si ya existe pero la variable cambió
#                     if not created and variable.variable != item['variable']:
#                         variable.variable = item['variable']
#                         variable.save()
#
#                     obj.variable_indicador.add(variable)
#
#             return True
#
#         except Exception as e:
#             messages.error(self.request, f"Error al guardar variables: {str(e)}")
#             return False
#
#     def _redirect_with_error(self, view_name, *args, error_message=None):
#         """Redirige con mensaje de error"""
#         if error_message:
#             messages.error(self.request, error_message)
#         return HttpResponseRedirect(reverse(view_name, args=args))
#
#     def handle_no_permission(self):
#         messages.error(self.request, 'Usted no tiene los privilegios necesarios para esta operación.')
#         return redirect(reverse('registro:lista_indicador', args=[self.id_accion]))


class FormulaValidatorInterface(ABC):
    """Interface para validadores de fórmulas (Dependency Inversion Principle)"""

    @abstractmethod
    def validate(self, formula: str, variables: List[str]) -> bool:
        pass

    @abstractmethod
    def get_variables_from_formula(self, formula: str) -> List[str]:
        pass


class RegexFormulaValidator(FormulaValidatorInterface):
    """Implementación concreta del validador de fórmulas"""

    def validate(self, formula: str, variables: List[str]) -> bool:
        """Valida que todas las variables de la fórmula estén definidas"""
        formula_variables = self.get_variables_from_formula(formula)
        return all(var in variables for var in formula_variables)

    def get_variables_from_formula(self, formula: str) -> List[str]:
        """Extrae variables de la fórmula usando expresión regular"""
        return re.findall(r'\b[a-zA-Z_][a-zA-Z0-9_]*\b', formula)


class VariableIndicadorService:
    """Servicio para manejar las variables de indicadores (Single Responsibility)"""

    def __init__(self):
        self.variable_model = None  # Se inyectará desde el view

    def set_model(self, model_class):
        """Inyección de dependencia del modelo"""
        self.variable_model = model_class

    def create_or_update_variables(self, json_data: List[Dict[str, Any]], indicador) -> bool:
        """Crea o actualiza variables de indicador"""
        try:
            for item in json_data:
                if item.get('nombre'):
                    variable, created = self.variable_model.objects.get_or_create(
                        nombre=item['nombre'],
                        defaults={'variable': item['variable']}
                    )

                    if not created and variable.variable != item['variable']:
                        variable.variable = item['variable']
                        variable.save()

                    indicador.variable_indicador.add(variable)
            return True
        except Exception:
            return False

    def extract_variables_from_json(self, json_data: List[Dict[str, Any]]) -> List[str]:
        """Extrae nombres de variables del JSON"""
        return [item['variable'] for item in json_data if 'variable' in item]


class IndicadorService:
    """Servicio principal para manejo de indicadores (Single Responsibility)"""

    def __init__(self, formula_validator: FormulaValidatorInterface, variable_service: VariableIndicadorService):
        self.formula_validator = formula_validator
        self.variable_service = variable_service

    def create_indicador(self, form, accion, json_data: List[Dict[str, Any]], request) -> bool:
        """Crea un indicador con sus variables asociadas"""
        try:
            with transaction.atomic():
                # Validar fórmula
                if not self._validate_formula(form.cleaned_data['formula'], json_data, request):
                    return False

                # Crear indicador
                indicador = form.save()
                accion.indicadores.add(indicador)

                # Crear variables asociadas
                if not self.variable_service.create_or_update_variables(json_data, indicador):
                    raise Exception("Error al crear variables")

                return True
        except Exception as e:
            messages.error(request, f'{str(e)} Ha ocurrido un error contacte con su administrador.')
            return False

    def update_indicador(self, form, json_data: List[Dict[str, Any]], request) -> bool:
        """Actualiza un indicador con sus variables asociadas"""
        try:
            with transaction.atomic():
                # Validar fórmula
                if not self._validate_formula(form.cleaned_data['formula'], json_data, request):
                    return False

                # Actualizar indicador
                indicador = form.save()
                indicador.variable_indicador.clear()

                # Actualizar variables asociadas
                if not self.variable_service.create_or_update_variables(json_data, indicador):
                    raise Exception("Error al actualizar variables")

                return True
        except Exception as e:
            messages.error(request, f'{str(e)} Ha ocurrido un error contacte con su administrador.')
            return False

    def _validate_formula(self, formula: str, json_data: List[Dict[str, Any]], request) -> bool:
        """Valida que la fórmula sea correcta"""
        variables = self.variable_service.extract_variables_from_json(json_data)

        if not self.formula_validator.validate(formula, variables):
            missing_variables = set(self.formula_validator.get_variables_from_formula(formula)) - set(variables)
            for var in missing_variables:
                messages.error(request,
                               f"La variable [{var}] no aparece en la fórmula del indicador. "
                               f"Por favor revise que las variables estén bien descritas.")
            return False
        return True


# ============================================================================
# BASE CLASSES - Evitando duplicación de código (DRY Principle)
# ============================================================================

class BaseIndicadorView:
    """Clase base para views de indicadores (evita duplicación)"""

    def __init__(self):
        self.accion = None
        self.id_accion = None

    def setup_accion(self, id_accion):
        """Configura la acción para el view"""
        self.id_accion = id_accion
        self.accion = get_object_or_404(Accion, id=id_accion)

    def get_cancel_url(self):
        """URL de cancelación común"""
        if self.accion.indicadores.all().count() > 0:
            return reverse('registro:lista_indicador', args=[self.id_accion])
        return reverse('registro:lista_accion')

    def handle_no_permission_redirect(self, request):
        """Manejo común de falta de permisos"""
        messages.error(request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(reverse('registro:lista_indicador', args=[self.id_accion]))


# ============================================================================
# VIEWS REFACTORIZADAS - Aplicando principios SOLID
# ============================================================================

class IndicadoresListView(LoginRequiredMixin, PermissionRequiredMixin, ListView, BaseIndicadorView):
    """Vista de lista de indicadores (Open/Closed Principle - extendible sin modificación)"""

    permission_required = 'registro.view_indicador'
    template_name = 'indicador/indicadores.html'
    context_object_name = 'indicadores'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseIndicadorView.__init__(self)

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)

    def get_queryset(self):
        self.setup_accion(self.kwargs['id_accion'])
        return self.accion.indicadores.all().order_by('nombre')

    def get(self, request, *args, **kwargs):
        if self.get_queryset().count() == 0:
            return HttpResponseRedirect(
                reverse('registro:registrar_indicador', args=[self.kwargs['id_accion']]))
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title_html': f'Listado indicadores de {self.accion.nombre}',
            'title_head': f'Indicadores de la acción [{self.kwargs["id_accion"]}]',
            'cobeneficio': True,
            'accion': self.accion,
            'crear_url': reverse('registro:registrar_indicador', args=[self.kwargs['id_accion']]),
            'show_menu_left': True,
            'breadcrumbs': BreadcrumbBuilder.build_indicador_list_breadcrumbs(),
            'current_step': 3,
            'next_url': '#'
        })
        return context


class IndicadorCreateView(LoginRequiredMixin, PermissionRequiredMixin, FormView, BaseIndicadorView):
    """Vista de creación de indicadores (Single Responsibility - solo creación)"""

    form_class = IndicadorForm
    template_name = 'indicador/crear_editar_indicador_accion.html'
    permission_required = 'registro.add_indicador'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseIndicadorView.__init__(self)
        # Inyección de dependencias
        formula_validator = RegexFormulaValidator()
        variable_service = VariableIndicadorService()
        variable_service.set_model(VariableIndicador)  # Inyección del modelo
        self.indicador_service = IndicadorService(formula_validator, variable_service)

    def dispatch(self, request, *args, **kwargs):
        self.setup_accion(kwargs.get('id_accion'))
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('registro:lista_indicador', args=[self.id_accion])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title_html': 'Indicador',
            'title_head': f'Indicador de la acción [{self.id_accion}]',
            'url_cancel': self.get_cancel_url(),
            'form_variable_indicador': VariableIndicadorForm(),
            'accion': self.accion,
            'show_menu_left': True,
            'indicadores': self.accion.indicadores.all(),
            'tipo_indicador': TipoIndicador.objects.all(),
            'breadcrumbs': BreadcrumbBuilder.build_indicador_create_breadcrumbs(self.id_accion),
            'current_step': 3.1,
            'next_url': '#'
        })
        return context

    def form_valid(self, form):
        json_data = json.loads(self.request.POST.get('form_variables', '[]'))

        if self.indicador_service.create_indicador(form, self.accion, json_data, self.request):
            messages.success(self.request, 'El indicador se ha registrado correctamente')
            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(reverse('registro:lista_indicador', args=[self.id_accion]))

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)


class IndicadorUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView, BaseIndicadorView):
    """Vista de actualización de indicadores (Single Responsibility - solo actualización)"""

    model = Indicador
    form_class = IndicadorForm
    template_name = 'indicador/crear_editar_indicador_accion.html'
    permission_required = 'registro.change_indicador'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseIndicadorView.__init__(self)
        # Inyección de dependencias
        formula_validator = RegexFormulaValidator()
        variable_service = VariableIndicadorService()
        variable_service.set_model(VariableIndicador)
        self.indicador_service = IndicadorService(formula_validator, variable_service)

    def dispatch(self, request, *args, **kwargs):
        self.setup_accion(kwargs.get('id_accion'))
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        id_indicador = self.kwargs.get('id_indicador')
        return get_object_or_404(Indicador, id=id_indicador)

    def get_success_url(self):
        return reverse('registro:lista_indicador', args=[self.id_accion])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        VariableIndicadorFormSet = modelformset_factory(
            VariableIndicador,
            form=VariableIndicadorForm,
            extra=0,
            can_delete=True
        )

        context.update({
            'title_html': 'Editar Indicador',
            'title_head': f'Indicador de la acción [{self.id_accion}]',
            'url_cancel': self.get_cancel_url(),
            'formset': VariableIndicadorFormSet(queryset=self.object.variable_indicador.all()),
            'accion': self.accion,
            'show_menu_left': True,
            'indicadores': self.accion.indicadores.all(),
            'tipo_indicador': TipoIndicador.objects.all(),
            'breadcrumbs': BreadcrumbBuilder.build_indicador_update_breadcrumbs(self.id_accion),
            'current_step': 3.1,
            'next_url': '#'
        })
        return context

    def form_valid(self, form):
        json_data = json.loads(self.request.POST.get('formset_variables', '[]'))

        if self.indicador_service.update_indicador(form, json_data, self.request):
            messages.success(self.request, 'El indicador se ha actualizado correctamente')
            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(reverse('registro:lista_indicador', args=[self.id_accion]))

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)


@csrf_exempt
@login_required
@permission_required('registro.delete_indicador', raise_exception=True)
def eliminar_indicador(request, id_accion, id_indicador):
    if request.method == 'POST':
        indicador = Indicador.objects.filter(id=id_indicador).first()
        indicador.variable_indicador.clear()
        indicador.resultados.all().delete()
        indicador.delete()
        messages.success(request, 'El indicador se ha eliminado correctamente')
        return HttpResponseRedirect(
            reverse('registro:lista_indicador', args=[id_accion]))


# ============================================================================
# BASE CLASSES - Evitando duplicación de código (DRY Principle)
# ============================================================================

class BaseResultadoIndicadorView:
    """Clase base para views de resultados de indicadores (evita duplicación)"""

    def __init__(self):
        self.accion = None
        self.indicador = None
        self.id_accion = None
        self.id_indicador = None

    def setup_accion_indicador(self, id_accion, id_indicador):
        """Configura la acción e indicador para el view"""
        self.id_accion = id_accion
        self.id_indicador = id_indicador
        self.accion = get_object_or_404(Accion, id=id_accion)
        self.indicador = get_object_or_404(Indicador, id=id_indicador)

    def get_list_url(self):
        """URL de lista común"""
        return reverse('registro:lista_resultado_indicador',
                       args=[self.id_accion, self.id_indicador])

    def get_cancel_url(self):
        """URL de cancelación común"""
        return reverse('registro:lista_indicador', args=[self.id_accion])

    def get_create_url(self):
        """URL de creación común"""
        return reverse('registro:registrar_resultado_indicador',
                       args=[self.id_accion, self.id_indicador])

    def handle_no_permission_redirect(self, request):
        """Manejo común de falta de permisos"""
        messages.error(request, 'Usted no tiene los privilegios necesarios para esta operación.')
        return redirect(self.get_list_url())


# ============================================================================
# VIEWS REFACTORIZADAS - Aplicando principios SOLID
# ============================================================================

class ResultadosIndicadorListView(LoginRequiredMixin, PermissionRequiredMixin,
                                  ListView, BaseResultadoIndicadorView):
    """Vista de lista de resultados de indicadores (Open/Closed Principle - extendible sin modificación)"""

    permission_required = 'registro.view_resultadoindicador'
    template_name = 'indicador/comportamiento.html'
    context_object_name = 'object_list'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseResultadoIndicadorView.__init__(self)
        # Inyección de dependencias
        self.variation_calculator = VariationCalculatorService()
        self.chart_data_service = ChartDataService()

        self.statistics_calculator = StatisticsCalculatorService()
        # self.projection_service = ProjectionService()
        self.insight_generator = InsightGeneratorService()

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)

    def dispatch(self, request, *args, **kwargs):
        self.setup_accion_indicador(kwargs.get('id_accion'), kwargs.get('id_indicador'))
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return self.indicador.resultados.all().order_by('fecha')

    def get(self, request, *args, **kwargs):
        """Redirige a crear si no hay resultados"""
        if self.get_queryset().count() == 0:
            return HttpResponseRedirect(self.get_create_url())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        object_list = self.get_queryset()

        # Usar servicios para calcular datos
        variations = self.variation_calculator.calculate_variations(object_list, self.indicador)
        chart_data = self.chart_data_service.prepare_chart_data(object_list, self.indicador)
        # Nuevos cálculos estadísticos
        advanced_stats = self.statistics_calculator.calculate_advanced_statistics(object_list, self.indicador)
        insights = self.insight_generator.generate_insights(object_list, advanced_stats, variations, self.indicador)

        # Cálculo de progreso hacia meta (si existe)
        meta_progress = self.indicador.calcular_progreso_meta()

        meta_progress_service = MetaProgressService()
        detailed_meta_progress = meta_progress_service.calculate_detailed_progress(self.indicador)

        # Próxima medición
        next_measurement = self._calculate_next_measurement()

        # RAnking
        todos_indicadores = Indicador.objects.all()
        performance_ranking = RankingCalculatorService.calculate_ranking(
            self.indicador, todos_indicadores
        )

        # Resumen ejecutivo y score de impacto climático
        executive_summary = self.insight_generator.generate_executive_summary(
            object_list, advanced_stats, variations, self.indicador
        )
        climate_impact_score = self.insight_generator.generate_climate_impact_score(
            object_list, advanced_stats, variations, self.indicador
        )

        # ============ NUEVOS GRÁFICOS MEJORADOS ============
        # 1. Gráfico de Tendencia con Regresión Lineal
        trend_regression_chart = self._prepare_trend_regression_chart(object_list)

        # 2. Gráfico de Cumplimiento de Meta vs Tiempo
        goal_timeline_chart = self._prepare_goal_timeline_chart(object_list)

        # 3. Gráfico de Comparación con otros indicadores de la acción
        comparison_chart = self._prepare_comparison_chart()

        # Obtener lista de indicadores de la misma acción para comparación
        indicadores_accion = self.accion.indicadores.exclude(id=self.indicador.id)

        context.update({
            'title_html': 'Resultados del indicador',
            'title_head': f'Resultados del indicador [{self.indicador.id}]',
            'resultado': True,
            'accion': self.accion,
            'indicador': self.indicador,
            'crear_url': self.get_create_url(),
            'url_cancel': self.get_cancel_url(),
            'primer_resultado': object_list.first(),
            'chart_data_line': chart_data,
            'dict_object_list': self.get_queryset(),
            'show_menu_left': True,
            'next_url': '#',
            'current_step': 3.2,
            'breadcrumbs': BreadcrumbBuilder.build_comportamiento_indicador_breadcrumbs(self.id_accion,
                                                                                        self.id_indicador),
            'insights': insights,
            'executive_summary': executive_summary,
            'climate_impact_score': climate_impact_score,
            'advanced_statistics': advanced_stats,
            'meta_progress': meta_progress,
            'detailed_meta_progress': detailed_meta_progress,
            'next_measurement': next_measurement,
            'performance_level': self._get_performance_level(variations, advanced_stats),
            'performance_ranking': performance_ranking,
            # Nuevos gráficos
            'trend_regression_chart': json.dumps(trend_regression_chart),
            'goal_timeline_chart': json.dumps(goal_timeline_chart),
            'comparison_chart': json.dumps(comparison_chart),
            'indicadores_accion': indicadores_accion,
            **variations
        })

        return context

    def _prepare_trend_regression_chart(self, object_list):
        """Gráfico de tendencia con línea de regresión y proyección"""
        if object_list.count() < 3:
            return None

        import numpy as np
        from datetime import timedelta

        # Datos reales
        fechas = []
        valores = []
        labels = []

        for idx, resultado in enumerate(object_list):
            fechas.append(idx)
            valores.append(round(float(resultado.valor),2))
            labels.append(resultado.fecha.strftime("%d-%m-%Y"))

        # Calcular regresión lineal
        x = np.array(fechas)
        y = np.array(valores)

        # Regresión lineal: y = mx + b
        m, b = np.polyfit(x, y, 1)

        # Línea de tendencia
        trend_line = [round(float(m * xi + b),2) for xi in x]

        # Proyección (próximos 3 puntos)
        future_x = list(range(len(x), len(x) + 3))
        future_trend = [round(float(m * xi + b),2) for xi in future_x]

        # Preparar etiquetas para proyección
        last_date = object_list.last().fecha
        future_labels = []
        for i in range(1, 4):
            future_date = last_date + timedelta(days=30 * i)
            future_labels.append(future_date.strftime("%d-%m-%Y"))

        # Calcular banda de confianza (±1 desviación estándar)
        residuals = y - np.array(trend_line)
        std_dev = round(float(np.std(residuals)),2)

        upper_band = [round(float(val + std_dev),2) for val in trend_line]
        lower_band = [round(float(val - std_dev),2) for val in trend_line]

        # IMPORTANTE: Convertir m y b a float nativo para el subtitle
        m_float = float(m)
        b_float = float(b)
        r_squared = float(self._calculate_r_squared(y, trend_line))

        chart_config = {
            'series': [
                {
                    'name': 'Valores Reales',
                    'type': 'scatter',
                    'data': valores
                },
                {
                    'name': 'Línea de Tendencia',
                    'type': 'line',
                    'data': trend_line
                },
                {
                    'name': 'Proyección',
                    'type': 'line',
                    'data': [None] * (len(valores) - 1) + [valores[-1]] + future_trend
                },
                {
                    'name': 'Banda Superior',
                    'type': 'line',
                    'data': upper_band
                },
                {
                    'name': 'Banda Inferior',
                    'type': 'line',
                    'data': lower_band
                }
            ],
            'chart': {
                'height': 400,
                'type': 'line',
                'toolbar': {'show': 'true'},
                'zoom': {'enabled': 'true'}
            },
            'title': {
                'text': '',
                'style': {
                    'fontSize': '18px',
                    'fontWeight': 'bold',
                    'fontFamily': 'Inter, sans-serif'
                }
            },
            'colors': ['#1B84FF', '#10b981', '#f59e0b', '#e0e0e0', '#e0e0e0'],
            'stroke': {
                'width': [0, 3, 3, 1, 1],
                'curve': ['smooth', 'straight', 'straight', 'smooth', 'smooth'],
                'dashArray': [0, 0, 5, 3, 3]
            },
            'markers': {
                'size': [8, 0, 0, 0, 0],
                'hover': {'size': [10, 0, 0, 0, 0]}
            },
            'fill': {
                'type': ['solid', 'solid', 'solid', 'solid', 'solid'],
                'opacity': [1, 1, 1, 0.1, 0.1]
            },
            'xaxis': {
                'categories': labels + future_labels,
                'labels': {
                    'rotate': -45,
                    'style': {'fontSize': '11px'}
                }
            },
            'yaxis': {
                'title': {
                    'text': self.indicador.unidad_medida.nombre if self.indicador.unidad_medida else 'Valor'
                }
            },
            'legend': {
                'position': 'top',
                'horizontalAlign': 'center'
            },
            'annotations': {
                'xaxis': [{
                    'x': len(labels) - 0.5,
                    'borderColor': '#999',
                    'strokeDashArray': 5,
                    'label': {
                        'text': 'Proyección →',
                        'style': {
                            'color': '#fff',
                            'background': '#f59e0b'
                        }
                    }
                }]
            }
        }

        # Agregar información de tendencia
        chart_config['subtitle'] = {
            'text': f'Tendencia: {"Creciente" if m > 0 else "Decreciente"} | Pendiente: {m:.4f} | R²: {self._calculate_r_squared(y, trend_line):.3f}',
            'style': {'fontSize': '12px', 'color': '#666'}
        }

        return chart_config

    def _prepare_goal_timeline_chart(self, object_list):
        """Gráfico de progreso hacia meta con timeline"""
        if not self.indicador.meta_valor or object_list.count() < 2:
            return None

        # Datos para el gráfico
        fechas = []
        valores = []
        progreso_porcentaje = []
        meta = round(self.indicador.meta_valor,2)
        baseline = round(self.indicador.valor_baseline or 0,2)

        for resultado in object_list:
            fechas.append(resultado.fecha.strftime("%d-%m-%Y"))
            valores.append(float(resultado.valor))

            # Calcular progreso según dirección
            if self.indicador.direccion_optima == 'incremento':
                if meta > baseline:
                    progreso = round(((resultado.valor - baseline) / (meta - baseline)) * 100,2)
                else:
                    progreso = 0
            else:  # decremento
                if baseline > meta:
                    progreso = round(((baseline - resultado.valor) / (baseline - meta)) * 100,2)
                else:
                    progreso = 0

            progreso_porcentaje.append(min(100, max(0, progreso)))

        # Determinar colores según progreso
        colors = []
        for prog in progreso_porcentaje:
            if prog >= 80:
                colors.append('#10b981')  # Verde
            elif prog >= 50:
                colors.append('#f59e0b')  # Amarillo
            else:
                colors.append('#ef4444')  # Rojo

        chart_config = {
            'series': [
                {
                    'name': 'Valor Actual',
                    'type': 'line',
                    'data': valores
                },
                {
                    'name': '% Progreso Meta',
                    'type': 'bar',
                    'data': progreso_porcentaje
                }
            ],
            'chart': {
                'height': 400,
                'type': 'line',
                'toolbar': {'show': 'true'}
            },
            'title': {
                'text': 'Evolución hacia la Meta Climática',
                'style': {
                    'fontSize': '18px',
                    'fontWeight': 'bold',
                    'fontFamily': 'Inter, sans-serif'
                }
            },
            'colors': ['#1B84FF', '#10b981'],
            'stroke': {
                'width': [3, 0],
                'curve': 'smooth'
            },
            'plotOptions': {
                'bar': {
                    'columnWidth': '50%',
                    'borderRadius': 4,
                    'distributed': 'false'
                }
            },
            'dataLabels': {
                'enabled': 'true',
                'enabledOnSeries': [1],
            },
            'xaxis': {
                'categories': fechas,
                'labels': {
                    'rotate': -45,
                    'style': {'fontSize': '11px'}
                }
            },
            'yaxis': [
                {
                    'title': {
                        'text': self.indicador.unidad_medida.nombre if self.indicador.unidad_medida else 'Valor'
                    },
                },
                {
                    'opposite': 'true',
                    'title': {
                        'text': '% Progreso'
                    },
                    'min': 0,
                    'max': 100,
                }
            ],
            'legend': {
                'position': 'top',
                'horizontalAlign': 'center'
            },
            'annotations': {
                'yaxis': [{
                    'y': meta,
                    'y2': None,
                    'borderColor': '#ef4444',
                    'strokeDashArray': 5,
                    'label': {
                        'text': f'Meta: {meta}',
                        'style': {
                            'color': '#fff',
                            'background': '#ef4444'
                        }
                    }
                }]
            }
        }

        return chart_config

    def _prepare_comparison_chart(self):
        """Gráfico comparativo con otros indicadores de la misma acción"""
        # Obtener indicadores de la misma acción
        indicadores_accion = self.accion.indicadores.all()

        if indicadores_accion.count() < 2:
            return None

        series_data = []

        for indicador in indicadores_accion:
            resultados = indicador.resultados.all().order_by('fecha')

            if resultados.count() < 2:
                continue

            # Normalizar valores para comparación (0-100)
            valores = [float(r.valor) for r in resultados]
            min_val = min(valores)
            max_val = max(valores)

            if max_val - min_val > 0:
                valores_normalizados = [
                    round(((val - min_val) / (max_val - min_val)) * 100,2)
                    for val in valores
                ]
            else:
                valores_normalizados = [50] * len(valores)

            # Tomar solo las últimas 10 mediciones para claridad
            if len(valores_normalizados) > 10:
                valores_normalizados = valores_normalizados[-10:]

            series_data.append({
                'name': indicador.nombre[:30] + '...' if len(indicador.nombre) > 30 else indicador.nombre,
                'data': valores_normalizados,
                'id': indicador.id
            })

        if not series_data:
            return None

        # Crear categorías genéricas
        max_length = max(len(s['data']) for s in series_data)
        categories = [f'M-{i + 1}' for i in range(max_length)]

        chart_config = {
            'series': series_data,
            'chart': {
                'height': 400,
                'type': 'radar',
                'toolbar': {'show': 'true'},
                'dropShadow': {
                    'enabled': 'true',
                    'blur': 1,
                    'left': 1,
                    'top': 1
                }
            },
            'title': {
                'text': f'Comparación de Indicadores - {self.accion.nombre}',
                'style': {
                    'fontSize': '18px',
                    'fontWeight': 'bold',
                    'fontFamily': 'Inter, sans-serif'
                }
            },
            'colors': ['#1B84FF', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#14b8a6'],
            'stroke': {
                'width': 2
            },
            'fill': {
                'opacity': 0.2
            },
            'markers': {
                'size': 4,
                'hover': {
                    'size': 6
                }
            },
            'xaxis': {
                'categories': categories
            },
            'yaxis': {
                'show': 'true',
                'min': 0,
                'max': 100,
                'tickAmount': 5
            },
            'legend': {
                'position': 'bottom',
                'horizontalAlign': 'center',
                'fontSize': '12px'
            },
            'tooltip': {
                'y': {
                    'formatter': "function(val) { return val.toFixed(1) + ' (normalizado)' }"
                }
            }
        }

        return chart_config


    def _calculate_r_squared(self, y_actual, y_predicted):
        """Calcula el coeficiente de determinación R²"""
        import numpy as np
        y_actual = np.array(y_actual)
        y_predicted = np.array(y_predicted)

        ss_res = np.sum((y_actual - y_predicted) ** 2)
        ss_tot = np.sum((y_actual - np.mean(y_actual)) ** 2)

        r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0

        # CONVERTIR A FLOAT NATIVO
        return float(r_squared)

    def _calculate_meta_progress(self, object_list):
        """Calcula el progreso hacia la meta si existe"""
        if not object_list.exists():
            return None

        # Asumiendo que tienes un campo meta en el indicador o acción
        meta_value = getattr(self.indicador, 'meta_valor', None) or getattr(self.accion, 'meta_numerica', None)

        if not meta_value:
            return None

        ultimo_valor = object_list.last().valor
        progreso_porcentaje = (ultimo_valor / meta_value) * 100

        return {
            'meta_valor': meta_value,
            'valor_actual': ultimo_valor,
            'progreso_porcentaje': min(progreso_porcentaje, 100),
            'falta_para_meta': max(0, meta_value - ultimo_valor),
            'superado': ultimo_valor >= meta_value
        }

    def _calculate_next_measurement(self):
        """Calcula cuándo debe ser la próxima medición"""
        proxima_fecha = self.indicador.calcular_proxima_medicion().astimezone(timezone.get_current_timezone())
        if not proxima_fecha:
            return None

        hoy = timezone.now().astimezone(timezone.get_current_timezone())
        if proxima_fecha:
            diferencia = proxima_fecha - hoy
            dias_restantes = diferencia.days
            return {
                'fecha': proxima_fecha,
                'dias_restantes': dias_restantes,
                'urgente': dias_restantes <= 7,
                'vencida': dias_restantes < 0,
                'frecuencia': self.indicador.frecuencia_medicion.nombre if self.indicador.frecuencia_medicion else 'No definida'
            }
        return None

    def _get_performance_level(self, variations, statistics):
        """Determina el nivel de rendimiento general"""
        score = 0

        # Puntuación por efectividad climática
        efectividad = statistics.get('efectividad', {}).get('nivel', 'insuficiente')
        if efectividad == 'alta':
            score += 4
        elif efectividad == 'media':
            score += 2
        elif efectividad == 'baja':
            score += 1

        # Puntuación por interpretación de cambio
        if variations.get('interpretacion_cambio_total') == 'positivo':
            score += 3
        elif variations.get('interpretacion_cambio_reciente') == 'positivo':
            score += 1

        # Puntuación por consistencia
        coef_var = statistics.get('coef_variacion', 100)
        if coef_var < 10:
            score += 2
        elif coef_var < 20:
            score += 1

        if score >= 7:
            return {'nivel': 'excelente', 'texto': 'Acción Climática Ejemplar', 'color': 'success'}
        elif score >= 5:
            return {'nivel': 'bueno', 'texto': 'Acción Climática Efectiva', 'color': 'primary'}
        elif score >= 3:
            return {'nivel': 'regular', 'texto': 'Acción en Desarrollo', 'color': 'warning'}
        else:
            return {'nivel': 'critico', 'texto': 'Requiere Intervención Urgente', 'color': 'danger'}


class ResultadoIndicadorCreateView(LoginRequiredMixin, PermissionRequiredMixin,
                                   FormView, BaseResultadoIndicadorView):
    """Vista de creación de resultados de indicadores (Single Responsibility - solo creación)"""

    form_class = ResultadoIndicadorForm
    template_name = 'indicador/resultado_indicador.html'
    permission_required = 'registro.add_resultadoindicador'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseResultadoIndicadorView.__init__(self)
        # Inyección de dependencias
        formula_calculator = FormulaCalculatorService()
        self.resultado_service = ResultadoIndicadorService(formula_calculator)

    def dispatch(self, request, *args, **kwargs):
        self.setup_accion_indicador(kwargs.get('id_accion'), kwargs.get('id_indicador'))
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return self.get_list_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        variables = self.indicador.variable_indicador.all()
        ResultadoVariablesFormSet = formset_factory(ResultadoVariableForm, extra=0)

        if 'formset_variables' not in context:
            context['formset_variables'] = ResultadoVariablesFormSet(
                initial=[
                    {
                        'variable_indicador': var.nombre,
                        'variable_indicador_id': var.id
                    }
                    for var in variables
                ]
            )

        context.update({
            'title_html': 'Resultado Indicador',
            'title_head': f'Resultado de [{self.indicador.nombre}]',
            'url_cancel': self.get_success_url(),
            'accion': self.accion,
            'indicador': self.indicador,
            'variables': variables,
            'show_menu_left': True,
            'next_url': '#',
            'current_step': 3.3,
            'tipo_indicador': TipoIndicador.objects.all(),
        })

        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form()

        ResultadoVariablesFormSet = formset_factory(ResultadoVariableForm, extra=0)
        formset_variables = ResultadoVariablesFormSet(request.POST)

        if form.is_valid() and formset_variables.is_valid():
            return self.form_valid(form, formset_variables)
        else:
            return self.form_invalid(form, formset_variables=formset_variables)

    def form_valid(self, form, formset_variables):
        try:
            with transaction.atomic():
                # Crear el objeto resultado
                resultado_obj = form.save()

                # Usar el servicio para guardar y calcular
                self.resultado_service.save_resultado_with_calculation(
                    form, formset_variables, self.indicador, resultado_obj
                )

                messages.success(self.request, 'El resultado se ha registrado correctamente')
                return HttpResponseRedirect(self.get_success_url())

        except Exception as e:
            messages.error(
                self.request,
                'Error: [' + str(e) + ']'
            )
            return self.form_invalid(form, formset_variables=formset_variables)

    def form_invalid(self, form, **kwargs):
        context = self.get_context_data(form=form, **kwargs)
        return self.render_to_response(context)

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)


class ResultadoIndicadorUpdateView(LoginRequiredMixin, PermissionRequiredMixin,
                                   UpdateView, BaseResultadoIndicadorView):
    """Vista de actualización de resultados de indicadores (Single Responsibility - solo actualización)"""

    model = ResultadoIndicador
    form_class = ResultadoIndicadorForm
    template_name = 'indicador/resultado_indicador.html'
    permission_required = 'registro.change_resultadoindicador'
    pk_url_kwarg = 'id_resultado'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        BaseResultadoIndicadorView.__init__(self)
        # Inyección de dependencias
        formula_calculator = FormulaCalculatorService()
        self.resultado_service = ResultadoIndicadorService(formula_calculator)

    def dispatch(self, request, *args, **kwargs):
        self.setup_accion_indicador(kwargs.get('id_accion'), kwargs.get('id_indicador'))
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return self.get_list_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ResultadoVariablesFormSet = formset_factory(ResultadoVariableForm, extra=0)

        if 'formset_variables' not in context:
            context['formset_variables'] = ResultadoVariablesFormSet(
                initial=[
                    {
                        'variable_indicador': var.variable_indicador.nombre,
                        'variable_indicador_id': var.variable_indicador.id,
                        'valor': var.valor
                    }
                    for var in self.object.resultadovariable_set.all()
                ]
            )

        context.update({
            'title_html': 'Editar Resultado Indicador',
            'title_head': f'Resultado de [{self.indicador.nombre}]',
            'url_cancel': self.get_success_url(),
            'accion': self.accion,
            'indicador': self.indicador,
            'resultado': self.object,
            'variables': self.indicador.variable_indicador.all(),
            'tipo_indicador': TipoIndicador.objects.all(),
            'show_menu_left': True,
            'next_url': '#',
            'current_step': 3.3,
        })

        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        ResultadoVariablesFormSet = formset_factory(ResultadoVariableForm, extra=0)
        formset_variables = ResultadoVariablesFormSet(request.POST)

        if form.is_valid() and formset_variables.is_valid():
            return self.form_valid(form, formset_variables)
        else:
            return self.form_invalid(form, formset_variables=formset_variables)

    def form_valid(self, form, formset_variables):
        try:
            with transaction.atomic():
                # Actualizar el objeto resultado
                resultado_obj = form.save(commit=False)

                # Usar el servicio para actualizar y recalcular
                self.resultado_service.save_resultado_with_calculation(
                    form, formset_variables, self.indicador, resultado_obj
                )

                messages.success(self.request, 'El resultado se ha actualizado correctamente')
                return HttpResponseRedirect(self.get_success_url())

        except Exception as e:
            messages.error(
                self.request,
                'Error: [' + str(e) + ']'
            )
            return self.form_invalid(form, formset_variables=formset_variables)

    def form_invalid(self, form, **kwargs):
        context = self.get_context_data(form=form, **kwargs)
        return self.render_to_response(context)

    def handle_no_permission(self):
        return self.handle_no_permission_redirect(self.request)


@csrf_exempt
@login_required
@permission_required('registro.delete_resultadoindicador', raise_exception=True)
def eliminar_resultado_indicador(request, id_accion, id_indicador, id_resultado):
    if request.method == 'POST':
        indicador = ResultadoIndicador.objects.filter(id=id_resultado).first()
        indicador.delete()
        messages.success(request, 'El resultado se ha eliminado correctamente')
        return HttpResponseRedirect(
            reverse('registro:lista_resultado_indicador', args=[id_accion, id_indicador]))


def mapa_cuba_leaflet(request):
    data = {
        'title_html': 'Mapa de acciones',
        'tipo_acciones': TipoAccion.objects.all().order_by('orden'),
        'sectores': Sector.objects.all().order_by('nombre'),
        'estados': EstadoAccion.objects.all().order_by('orden'),
        'escenarios': Escenario.objects.all().order_by('nombre'),
    }
    return render(request, 'action/mapa.html', data)


@require_GET
def municipios_por_tipo_accion(request):
    tipo_id = request.GET.get('tipo')
    estado_id = request.GET.get('estado')
    sector_id = request.GET.get('sector')
    escenario_id = request.GET.get('escenario')
    fecha = request.GET.get('fecha')
    q = Q()
    # acciones = Accion.objects.filter(tipo_accion_id=tipo_id)
    # filtra con q estado, tipo, sector, escenario y fecha la fecha viene en formato 09/09/2025 to 18/09/2025
    if tipo_id:
        q &= Q(tipo_accion_id=tipo_id)
    if estado_id:
        q &= Q(estado_id=estado_id)
    if sector_id:
        q &= Q(sector_id=sector_id)
    if escenario_id:
        q &= Q(escenario_id=escenario_id)
    if fecha:
        try:
            fecha_inicio_str, fecha_fin_str = fecha.split(' to ')
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%d/%m/%Y').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%d/%m/%Y').date()
            q &= Q(fecha_inicio__gte=fecha_inicio) & Q(fecha_fin__lte=fecha_fin)
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido. Use DD/MM/YYYY to DD/MM/YYYY'}, status=400)

    acciones = Accion.objects.filter(q)
    municipios = {}
    provincias = {}
    tiene_municipios = False
    tiene_provincias = False

    for accion in acciones:
        # Si la acción tiene municipios
        if accion.municipios.exists():
            tiene_municipios = True
            for municipio in accion.municipios.all():
                prov = municipio.provincia
                if municipio.id not in municipios:
                    municipios[municipio.id] = {
                        'municipio_id': municipio.id,
                        'municipio_nombre': municipio.nombre,
                        'provincia_id': prov.id,
                        'provincia_nombre': prov.nombre,
                        'provincia_hc_keys': prov.hc_keys.lower(),
                        'acciones_count': 0,
                        'presupuesto_total': {},  # Cambiar a dict para sumar por moneda
                    }
                municipios[municipio.id]['acciones_count'] += 1
                for pp in accion.presupuestos_planificados.all():
                    moneda = pp.tipo_moneda.nombre
                    if moneda not in municipios[municipio.id]['presupuesto_total']:
                        municipios[municipio.id]['presupuesto_total'][moneda] = 0
                    municipios[municipio.id]['presupuesto_total'][moneda] += pp.monto
        # Si la acción tiene provincias
        elif accion.provincias.exists():
            tiene_provincias = True
            for provincia in accion.provincias.all():
                if provincia.id not in provincias:
                    provincias[provincia.id] = {
                        'provincia_id': provincia.id,
                        'provincia_nombre': provincia.nombre,
                        'provincia_hc_keys': provincia.hc_keys.lower(),
                        'acciones_count': 0,
                        'presupuesto_total': {},
                    }
                provincias[provincia.id]['acciones_count'] += 1
                for pp in accion.presupuestos_planificados.all():
                    moneda = pp.tipo_moneda.nombre
                    if moneda not in provincias[provincia.id]['presupuesto_total']:
                        provincias[provincia.id]['presupuesto_total'][moneda] = 0
                    provincias[provincia.id]['presupuesto_total'][moneda] += pp.monto

    # Convertir los dicts de presupuesto_total a listas de objetos {moneda, monto_total}
    for m in municipios.values():
        m['presupuesto_total'] = [{'moneda': k, 'monto_total': v} for k, v in m['presupuesto_total'].items() if v > 0]
    for p in provincias.values():
        p['presupuesto_total'] = [{'moneda': k, 'monto_total': v} for k, v in p['presupuesto_total'].items() if v > 0]

    # Calcular resumen ejecutivo
    total_acciones = acciones.count()
    total_activas = total_acciones  # Si tienes un campo para activas, cámbialo aquí
    resumen_presupuesto = {}
    for accion in acciones:
        for p in accion.presupuesto_total:
            moneda = p['moneda']
            monto = p['monto_total']
            if moneda not in resumen_presupuesto:
                resumen_presupuesto[moneda] = 0
            resumen_presupuesto[moneda] += monto
    resumen_presupuesto_list = [{'moneda': k, 'monto_total': v} for k, v in resumen_presupuesto.items() if v > 0]

    if municipios:
        print(resumen_presupuesto_list)
        return JsonResponse({
            'tipo': 'municipio',
            'municipios': list(municipios.values()),
            'resumen': {
                'total_acciones': total_acciones,
                'total_activas': total_activas,
                'presupuesto_total': resumen_presupuesto_list
            }
        }, safe=False)
    elif provincias:
        print(resumen_presupuesto_list)
        return JsonResponse({
            'tipo': 'provincia',
            'provincias': list(provincias.values()),
            'resumen': {
                'total_acciones': total_acciones,
                'total_activas': total_activas,
                'presupuesto_total': resumen_presupuesto_list
            }
        }, safe=False)
    else:
        return JsonResponse({'tipo': 'vacio', 'data': [],
                             'resumen': {'total_acciones': 0, 'total_activas': 0, 'presupuesto_total': []}}, safe=False)


@csrf_exempt
@login_required
@permission_required('registro.view_indicador', raise_exception=True)
def comparar_indicadores_accion(request, id_accion):
    """Vista para comparación dinámica de indicadores"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body)
        indicadores_ids = data.get('indicadores', [])

        if not indicadores_ids:
            return JsonResponse({'success': False, 'error': 'No se seleccionaron indicadores'})

        accion = get_object_or_404(Accion, id=id_accion)

        # Verificar que los indicadores pertenecen a la acción
        indicadores = accion.indicadores.filter(id__in=indicadores_ids)

        if not indicadores.exists():
            return JsonResponse({'success': False, 'error': 'Indicadores no encontrados'})

        # Preparar datos para el gráfico
        series_data = []

        for indicador in indicadores:
            resultados = indicador.resultados.all().order_by('fecha')

            if resultados.count() < 2:
                continue

            # Normalizar valores (0-100)
            valores = [float(r.valor) for r in resultados]
            min_val = min(valores)
            max_val = max(valores)

            if max_val - min_val > 0:
                valores_normalizados = [
                    ((val - min_val) / (max_val - min_val)) * 100
                    for val in valores
                ]
            else:
                valores_normalizados = [50] * len(valores)

            # Limitar a últimas 10 mediciones
            if len(valores_normalizados) > 10:
                valores_normalizados = valores_normalizados[-10:]

            series_data.append({
                'name': indicador.nombre[:30] + '...' if len(indicador.nombre) > 30 else indicador.nombre,
                'data': valores_normalizados,
                'id': indicador.id
            })

        return JsonResponse({
            'success': True,
            'series': series_data,
            'total': len(series_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('registro.view_indicador', raise_exception=True)
def exportar_indicador_excel(request, id_accion, id_indicador):
    """Exporta los datos del indicador a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    indicador = get_object_or_404(Indicador, id=id_indicador)
    accion = get_object_or_404(Accion, id=id_accion)
    resultados = indicador.resultados.all().order_by('fecha')

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Indicador"

    # Estilos
    header_fill = PatternFill(start_color="1B84FF", end_color="1B84FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Información del indicador
    ws['A1'] = 'REPORTE DE INDICADOR CLIMÁTICO'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    ws['A3'] = 'Acción:'
    ws['B3'] = accion.nombre
    ws['A4'] = 'Indicador:'
    ws['B4'] = indicador.nombre
    ws['A5'] = 'Tipo:'
    ws['B5'] = indicador.tipo_indicador.nombre
    ws['A6'] = 'Unidad de Medida:'
    ws['B6'] = indicador.unidad_medida.nombre if indicador.unidad_medida else 'N/A'
    ws['A7'] = 'Fecha de Generación:'
    ws['B7'] = datetime.now().strftime('%d/%m/%Y %H:%M')

    # Hacer negrita las etiquetas
    for row in range(3, 8):
        ws[f'A{row}'].font = Font(bold=True)

    # Encabezados de datos
    start_row = 9
    headers = ['Fecha', 'Valor', 'Fuente', 'Observación']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    for idx, resultado in enumerate(resultados, start=start_row + 1):
        ws.cell(row=idx, column=1).value = resultado.fecha.strftime('%d/%m/%Y')
        ws.cell(row=idx, column=2).value = resultado.valor
        ws.cell(row=idx, column=3).value = resultado.fuente_dato or 'N/A'
        ws.cell(row=idx, column=4).value = resultado.observacion or 'N/A'

        # Aplicar bordes
        for col in range(1, 5):
            ws.cell(row=idx, column=col).border = border

    # Estadísticas
    stats_row = len(resultados) + start_row + 2
    ws[f'A{stats_row}'] = 'ESTADÍSTICAS'
    ws[f'A{stats_row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{stats_row}:D{stats_row}')

    stats_row += 2

    if resultados.exists():
        valores = [r.valor for r in resultados]

        stats = [
            ('Valor Promedio:', sum(valores) / len(valores)),
            ('Valor Máximo:', max(valores)),
            ('Valor Mínimo:', min(valores)),
            ('Total Mediciones:', len(valores)),
            ('Primera Medición:', resultados.first().fecha.strftime('%d/%m/%Y')),
            ('Última Medición:', resultados.last().fecha.strftime('%d/%m/%Y')),
        ]

        if indicador.meta_valor:
            progreso = indicador.calcular_progreso_meta()
            if progreso:
                stats.append(('Meta:', indicador.meta_valor))
                stats.append(('Progreso hacia Meta:', f"{progreso['progreso_porcentaje']:.1f}%"))

        for idx, (label, value) in enumerate(stats):
            ws[f'A{stats_row + idx}'] = label
            ws[f'A{stats_row + idx}'].font = Font(bold=True)
            ws[f'B{stats_row + idx}'] = value

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'indicador_{indicador.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


class PresupuestoAnalyticsView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Vista para análisis presupuestario avanzado"""
    template_name = 'presupuesto/dashboard.html'
    permission_required = 'registro.view_presupuestoplanificado'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accion = get_object_or_404(Accion, id=self.kwargs['id_accion'])

        # Servicios
        analytics_service = PresupuestoAnalyticsService()
        forecast_service = PresupuestoForecastService()
        viz_service = PresupuestoVisualizationService()
        benchmark_service = PresupuestoBenchmarkService()
        indicador_analyzer = PresupuestoIndicadorAnalyzer()

        # Cálculos principales
        budget_health_score = analytics_service.calculate_budget_health_score(accion)
        alertas = analytics_service.detect_budget_anomalies(accion)
        resumen_ejecutivo = analytics_service.calculate_execution_summary(accion)
        proyeccion = forecast_service.forecast_budget_completion(accion)
        recomendaciones = forecast_service.recommend_budget_adjustments(accion)

        # Datos para gráficos
        waterfall_data = viz_service.generate_waterfall_data(accion)
        category_data = viz_service.generate_category_chart_data(accion)
        burn_rate_data = viz_service.generate_burn_rate_data(accion)

        # Benchmarking
        benchmarking = benchmark_service.compare_budget_efficiency(accion)

        # Costo-efectividad
        costo_efectividad = indicador_analyzer.calculate_cost_effectiveness(accion)

        # Breadcrumbs
        breadcrumbs = [
            {'name': 'Inicio', 'url': reverse('registro:home'), 'icon': 'ki-home'},
            {'name': 'Acciones', 'url': reverse('registro:lista_accion'), 'icon': None},
            {'name': 'Presupuesto', 'url': reverse('registro:lista_presupuesto_planificado', args=[accion.id]),
             'icon': None},
            {'name': 'Analytics', 'url': None, 'icon': None, 'active': True},
        ]

        context.update({
            'title_html': 'Análisis Presupuestario Avanzado',
            'title_head': f'Analytics - Acción [{accion.id}]',
            'accion': accion,
            'budget_health_score': budget_health_score,
            'alertas_presupuestarias': alertas,
            'resumen_ejecutivo': resumen_ejecutivo,
            'proyeccion_finalizacion': proyeccion,
            'recomendaciones': recomendaciones,
            'waterfall_data': waterfall_data,
            'category_data': category_data,
            'burn_rate_data': burn_rate_data,
            'benchmarking': benchmarking,
            'costo_efectividad': costo_efectividad,
            'presupuestos_planificados': accion.presupuestos_planificados.all(),
            'tipos_moneda': TipoMoneda.objects.filter(estado=True),
            'crear_url': reverse('registro:registrar_presupuesto_planificado', args=[accion.id]),
            'breadcrumbs': breadcrumbs,
            'show_menu_left': True,
            'current_step': 2
        })

        return context


@login_required
@require_GET
def burn_rate_data_ajax(request, id_accion):
    """Endpoint AJAX para datos de burn rate por moneda"""
    accion = get_object_or_404(Accion, id=id_accion)
    moneda_id = request.GET.get('moneda_id')

    moneda = None
    if moneda_id:
        moneda = get_object_or_404(TipoMoneda, id=moneda_id)

    viz_service = PresupuestoVisualizationService()
    data = viz_service.generate_burn_rate_data(accion, moneda)

    return JsonResponse(data)


@login_required
@permission_required('registro.view_presupuestoplanificado')
def exportar_reporte_presupuesto(request, id_accion):
    """Exporta reporte presupuestario en diferentes formatos"""
    import json
    from django.http import HttpResponse
    from io import BytesIO

    accion = get_object_or_404(Accion, id=id_accion)
    formato = request.GET.get('formato', 'pdf')
    opciones_json = request.GET.get('opciones', '{}')
    opciones = json.loads(opciones_json)

    # Servicios
    analytics_service = PresupuestoAnalyticsService()
    forecast_service = PresupuestoForecastService()

    # Recopilar datos
    datos_reporte = {
        'accion': accion,
        'fecha_generacion': timezone.now(),
        'usuario': request.user
    }

    if opciones.get('resumen'):
        datos_reporte['resumen'] = analytics_service.calculate_execution_summary(accion)
        datos_reporte['health_score'] = analytics_service.calculate_budget_health_score(accion)

    if opciones.get('alertas'):
        datos_reporte['alertas'] = analytics_service.detect_budget_anomalies(accion)

    if opciones.get('proyecciones'):
        datos_reporte['proyecciones'] = forecast_service.forecast_budget_completion(accion)
        datos_reporte['recomendaciones'] = forecast_service.recommend_budget_adjustments(accion)

    if formato == 'pdf':
        # Generar PDF
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1B84FF'),
            spaceAfter=30,
        )
        elements.append(Paragraph(f"Reporte Presupuestario - {accion.nombre}", title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Resumen Ejecutivo
        if opciones.get('resumen'):
            elements.append(Paragraph("Resumen Ejecutivo", styles['Heading2']))
            elements.append(Spacer(1, 0.1 * inch))

            resumen_data = [
                ['Métrica', 'Valor']
            ]

            for item in datos_reporte['resumen']['total_planificado']:
                resumen_data.append([
                    f"Total Planificado ({item['moneda']})",
                    f"{item['moneda_simbolo']}{item['monto_total']:,.2f}"
                ])

            for item in datos_reporte['resumen']['total_ejecutado']:
                resumen_data.append([
                    f"Total Ejecutado ({item['moneda']})",
                    f"{item['moneda_simbolo']}{item['monto_total']:,.2f} ({item['porcentaje']:.1f}%)"
                ])

            resumen_data.append([
                'Score de Salud',
                f"{datos_reporte['health_score']}/100"
            ])

            resumen_table = Table(resumen_data, colWidths=[3 * inch, 2 * inch])
            resumen_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B84FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(resumen_table)
            elements.append(Spacer(1, 0.3 * inch))

        # Alertas
        if opciones.get('alertas') and datos_reporte.get('alertas'):
            elements.append(Paragraph("Alertas y Notificaciones", styles['Heading2']))
            elements.append(Spacer(1, 0.1 * inch))

            for alerta in datos_reporte['alertas']:
                alerta_text = f"<b>{alerta['titulo']}</b><br/>{alerta['mensaje']}"
                elements.append(Paragraph(alerta_text, styles['Normal']))
                elements.append(Spacer(1, 0.1 * inch))

            elements.append(Spacer(1, 0.2 * inch))

        # Proyecciones
        if opciones.get('proyecciones') and datos_reporte.get('proyecciones'):
            elements.append(Paragraph("Proyecciones", styles['Heading2']))
            proyeccion = datos_reporte['proyecciones']

            proj_text = f"""
            Fecha estimada de finalización: <b>{proyeccion['fecha_estimada'].strftime('%d/%m/%Y')}</b><br/>
            Meses estimados: <b>{proyeccion['meses_estimados']}</b><br/>
            Tendencia: <b>{proyeccion['tendencia'].title()}</b>
            """
            elements.append(Paragraph(proj_text, styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))

        # Construir PDF
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_presupuesto_accion_{accion.id}.pdf"'
        return response

    elif formato == 'excel':
        # Generar Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Presupuestario"

        # Encabezado
        ws['A1'] = f"Reporte Presupuestario - {accion.nombre}"
        ws['A1'].font = Font(size=16, bold=True, color="1B84FF")
        ws.merge_cells('A1:D1')

        row = 3

        # Resumen
        if opciones.get('resumen'):
            ws[f'A{row}'] = "Resumen Ejecutivo"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            ws[f'A{row}'] = "Métrica"
            ws[f'B{row}'] = "Valor"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            for item in datos_reporte['resumen']['total_planificado']:
                ws[f'A{row}'] = f"Total Planificado ({item['moneda']})"
                ws[f'B{row}'] = item['monto_total']
                ws[f'B{row}'].number_format = '"$"#,##0.00'
                row += 1

            row += 2

        # Presupuestos detallados
        if opciones.get('detalle'):
            ws[f'A{row}'] = "Detalle de Presupuestos"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            headers = ['Tipo', 'Fuente', 'Planificado', 'Ejecutado', '% Ejecución', 'Restante']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1B84FF", end_color="1B84FF", fill_type="solid")
            row += 1

            for pp in accion.presupuestos_planificados.all():
                ws[f'A{row}'] = pp.tipo_presupuesto.nombre
                ws[f'B{row}'] = pp.fuente_financiamiento
                ws[f'C{row}'] = pp.monto
                ws[f'D{row}'] = pp.get_monto_total_ejecutado
                ws[f'E{row}'] = pp.get_porcentaje_monto_ejecutado / 100
                ws[f'F{row}'] = pp.get_monto_restante

                ws[f'C{row}'].number_format = '"$"#,##0.00'
                ws[f'D{row}'].number_format = '"$"#,##0.00'
                ws[f'E{row}'].number_format = '0.00%'
                ws[f'F{row}'].number_format = '"$"#,##0.00'

                row += 1

        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_presupuesto_accion_{accion.id}.xlsx"'
        return response

    elif formato == 'csv':
        # Generar CSV
        import csv

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="reporte_presupuesto_accion_{accion.id}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Tipo', 'Fuente', 'Moneda', 'Planificado', 'Ejecutado', 'Porcentaje', 'Restante'])

        for pp in accion.presupuestos_planificados.all():
            writer.writerow([
                pp.tipo_presupuesto.nombre,
                pp.fuente_financiamiento,
                pp.tipo_moneda.nombre,
                pp.monto,
                pp.get_monto_total_ejecutado,
                f"{pp.get_porcentaje_monto_ejecutado:.2f}%",
                pp.get_monto_restante
            ])

        return response

    return HttpResponse("Formato no soportado", status=400)


@login_required
@permission_required('registro.view_indicador', raise_exception=True)
def generar_reporte_pdf(request, id_accion, id_indicador, tipo_reporte):
    """Vista para generar diferentes tipos de reportes PDF"""

    indicador = get_object_or_404(Indicador, id=id_indicador)
    accion = get_object_or_404(Accion, id=id_accion)

    # Crear servicio de reportes
    reporte_service = ReportePDFService(indicador, accion)

    # Generar según tipo
    if tipo_reporte == 'completo':
        pdf_content = reporte_service.generar_reporte_completo()
        filename = f'reporte_completo_{indicador.id}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
    elif tipo_reporte == 'ejecutivo':
        pdf_content = reporte_service.generar_reporte_ejecutivo()
        filename = f'reporte_ejecutivo_{indicador.id}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
    elif tipo_reporte == 'estadistico':
        pdf_content = reporte_service.generar_reporte_estadistico()
        filename = f'reporte_estadistico_{indicador.id}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
    elif tipo_reporte == 'comparativo':
        pdf_content = reporte_service.generar_reporte_comparativo()
        filename = f'reporte_comparativo_{indicador.id}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
    else:
        return HttpResponse('Tipo de reporte no válido', status=400)

    # Preparar respuesta
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
@permission_required('registro.view_accion', raise_exception=True)
def generar_reporte_accion_completo(request, id_accion):
    """Genera reporte completo de toda la acción con todos sus indicadores"""
    accion = get_object_or_404(Accion, id=id_accion)
    indicadores = accion.indicadores.all()

    # Preparar contexto
    context = {
        'accion': accion,
        'indicadores': [],
        'fecha_generacion': datetime.datetime.now(),
    }

    # Generar datos para cada indicador
    for indicador in indicadores:
        resultados = indicador.resultados.all().order_by('fecha')
        if resultados.exists():
            statistics_calculator = StatisticsCalculatorService()
            stats = statistics_calculator.calculate_advanced_statistics(resultados, indicador)

            reporte_service = ReportePDFService(indicador, accion)

            context['indicadores'].append({
                'indicador': indicador,
                'estadisticas': stats,
                'ultimo_valor': resultados.last().valor,
                'total_mediciones': resultados.count(),
                'grafico': reporte_service._generar_grafico_tendencia(),
            })

    # Renderizar template
    html_string = render_to_string('reportes/reporte_accion_completa.html', context)

    # Generar PDF
    font_config = FontConfiguration()
    css = CSS(string='''
        @page { size: A4; margin: 2cm; }
        body { font-family: Arial, sans-serif; font-size: 10pt; }
        h1 { color: #1B84FF; page-break-before: always; }
        table { width: 100%; border-collapse: collapse; }
        th { background: #1B84FF; color: white; padding: 8pt; }
        td { border: 1pt solid #ddd; padding: 6pt; }
    ''', font_config=font_config)

    pdf_content = HTML(string=html_string).write_pdf(stylesheets=[css], font_config=font_config)

    # Respuesta
    response = HttpResponse(pdf_content, content_type='application/pdf')
    filename = f'reporte_accion_{accion.id}_{datetime.now().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response